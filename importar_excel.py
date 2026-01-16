"""
Script para importar servicios desde el Excel existente a la base de datos
"""
import sqlite3
from openpyxl import load_workbook
import os

def importar_desde_excel(excel_path, username='admin', password='admin123'):
    """
    Importa servicios desde el Excel a la base de datos
    """
    # Crear la base de datos si no existe
    from werkzeug.security import generate_password_hash
    
    db_path = 'database/gastos.db'
    os.makedirs('database', exist_ok=True)
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Crear tablas si no existen
    cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        email TEXT,
        telefono TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS servicios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        nombre TEXT NOT NULL,
        dia_vencimiento INTEGER,
        monto REAL,
        medio_pago TEXT,
        activo INTEGER DEFAULT 1,
        FOREIGN KEY (user_id) REFERENCES usuarios (id)
    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS pagos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        servicio_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        periodo TEXT NOT NULL,
        monto REAL NOT NULL,
        fecha_pago TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        metodo_pago TEXT,
        FOREIGN KEY (servicio_id) REFERENCES servicios (id),
        FOREIGN KEY (user_id) REFERENCES usuarios (id)
    )''')
    
    conn.commit()
    
    # Crear usuario admin si no existe
    cursor.execute('SELECT id FROM usuarios WHERE username = ?', (username,))
    user = cursor.fetchone()
    
    if not user:
        hashed_password = generate_password_hash(password)
        cursor.execute('INSERT INTO usuarios (username, password) VALUES (?, ?)',
                      (username, hashed_password))
        conn.commit()
        user_id = cursor.lastrowid
        print(f"✓ Usuario '{username}' creado con contraseña '{password}'")
    else:
        user_id = user[0]
        print(f"✓ Usuario '{username}' ya existe")
    
    # Cargar Excel
    print(f"\nLeyendo archivo Excel: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active
    
    servicios_importados = 0
    
    # Leer servicios desde la fila 5 hasta encontrar TOTAL MES
    for row_idx in range(5, 100):
        nombre = ws.cell(row=row_idx, column=1).value
        
        # Parar cuando lleguemos a TOTAL MES o celda vacía
        if not nombre or nombre == 'TOTAL MES':
            break
        
        # Saltar filas vacías
        if nombre.strip() == '':
            continue
        
        dia_vencimiento = ws.cell(row=row_idx, column=2).value
        monto = ws.cell(row=row_idx, column=3).value
        medio_pago_col = 6 if ws.max_column >= 6 else None
        medio_pago = ws.cell(row=row_idx, column=medio_pago_col).value if medio_pago_col else None
        
        # Convertir tipos
        try:
            dia_vencimiento = int(dia_vencimiento) if dia_vencimiento else None
        except:
            dia_vencimiento = None
        
        try:
            monto = float(monto) if monto else None
        except:
            monto = None
        
        # Verificar si el servicio ya existe
        cursor.execute('SELECT id FROM servicios WHERE user_id = ? AND nombre = ?',
                      (user_id, nombre))
        existe = cursor.fetchone()
        
        if not existe:
            cursor.execute('''
                INSERT INTO servicios (user_id, nombre, dia_vencimiento, monto, medio_pago)
                VALUES (?, ?, ?, ?, ?)
            ''', (user_id, nombre, dia_vencimiento, monto, medio_pago))
            servicios_importados += 1
            print(f"  ✓ Importado: {nombre}")
        else:
            print(f"  → Ya existe: {nombre}")
    
    conn.commit()
    conn.close()
    
    print(f"\n{'='*60}")
    print(f"✓ Importación completada!")
    print(f"  - Servicios nuevos importados: {servicios_importados}")
    print(f"  - Usuario: {username}")
    print(f"  - Contraseña: {password}")
    print(f"{'='*60}\n")
    print("Ahora podés iniciar la aplicación con: python app.py")
    print(f"Y loguearte con usuario '{username}' y contraseña '{password}'")

if __name__ == '__main__':
    import sys
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        # Buscar archivo Excel en el directorio actual
        excel_file = 'billetera_mata_galan.xlsx'

        if not os.path.exists(excel_file):
            print("Error: No se encontró el archivo Excel")
            print("\nUso: python importar_excel.py [ruta_al_excel.xlsx]")
            print("\nO copiá tu archivo Excel como 'billetera_mata_galan.xlsx' en esta carpeta")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"Error: No se encuentra el archivo {excel_file}")
        sys.exit(1)
    
    print("="*60)
    print("IMPORTADOR DE DATOS DESDE EXCEL")
    print("="*60)
    
    importar_desde_excel(excel_file)
